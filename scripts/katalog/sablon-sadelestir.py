# -*- coding: utf-8 -*-
"""
FİYAT ŞABLONUNU KALANLA SINIRLA (2026-08-28).

sablon-uret.py 658 satırlık tam şablonu üretiyordu. O günden bu yana Viniltürk
listesinden 7 ürün + eksik malzemeler + ek işlemler siteye girildi; artık 457
satırın karşılığı canlıda var ve Hasan'ın onlara bakması gereksiz.

Bu script CANLI SİTEYİ sorgular ve yalnız gerçekten eksik olanı bırakır:
  · fiyatı olan ürünün satırları atılır
  · matris ürünlerde YALNIZ EN DÜŞÜK kademe bırakılır — üst kademeler Hasan'ın
    kendi mevcut fiyatlarından çıkarılan katsayılarla türetilir (kartvizit ×1,93 /
    ×2,90 / ×4,79 / ×9,67 · broşür ×1,53 / ×2,88 / ×4,63 · çanta ×1,43)
  · EKSTRALAR sayfası tamamen düşer (CNC, laminasyon, dikiş… siteye girildi)

Sonuç: 658 satır → doldurulacak ~97 hücre.
"""
import json
import subprocess
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding="utf-8")
KAYNAK = r"C:\Users\Administrator\Downloads\markala-fiyat-girisi.xlsx"
HEDEF = r"C:\Users\Administrator\Downloads\markala-fiyat-girisi.xlsx"
API = "https://api.markala.com.tr/api/products?take=5000&list=true"

# Excel ürün adı → site slug. Sitede fiyatı VARSA o ürünün satırları atılır.
ESLE = {
    "Standart Kartvizit": "klasik-kartvizit", "Sıvama Kartvizit": "klasik-kartvizit",
    "Kabartma Laklı Kartvizit": "klasik-kartvizit", "Broşür": "brosur", "El İlanı": "el-ilani",
    "Kapı Askısı El İlanı": "kapi-aski-brosur", "Antetli Kağıt": "antetli-kagit",
    "Diplomat Zarf": "zarf-diplomat-renkli", "Cepli Dosya": "cepli-dosya", "Makbuz": "makbuz",
    "Küp Bloknot": "kup-bloknot", "Spiralli Bloknot/Defter": "spiralli-bloknot",
    "Amerikan Bristol Karton Çanta": "canta", "Kraft Karton Çanta": "canta",
    "Kağıt Oto Paspas": "oto-paspas", "Amerikan Servis": "amerikan-servis", "Sticker Baskı": "etiket",
    "Folyo Baskı": "folyo-cesitleri", "Avrupa Vinil Baskı": "vinil-branda-440gr",
    "Çin (Lamine) Vinil Baskı": "vinil-branda-440gr", "Mesh Delikli Vinil Baskı": "mesh-branda",
    "One Way Vision Baskı": "one-way-vision-baski", "Dekota\\Foreks Baskı": "dekota-baski-5mm",
    "Roll Up Banner": "rollup-standart", "Yelken Bayrak": "yelken-bayrak-damla",
    "Kırlangıç Bayrak": "kirlangic-bayrak-3m", "Makam Bayrağı": "makam-bayragi-puskullu",
    "Masa Bayrağı": "masa-bayragi-krom", "Baskes Folyo": "baskes-folyo",
    "Duvar Kağıdı": "duvar-kagidi-baski", "Pleksi Baskı": "pleksi-baski",
    "Kompozit Baskı": "kompozit-baski", "Kanvas Tablo Baskı": "kanvas-tablo-baski",
    "UV DTF Baskı": "uv-dtf-baski",
}

r = subprocess.run(["curl", "-s", API], capture_output=True, text=True, encoding="utf-8")
fiyatli = {p["slug"] for p in json.loads(r.stdout) if (p.get("displayPrice") or 0) > 0}
print(f"sitede fiyatı olan ürün: {len(fiyatli)}")

wb_in = openpyxl.load_workbook(KAYNAK)
src = wb_in["FİYAT GİRİŞİ"]
basliklar = [c.value for c in src[1]]

# Satırları topla; matris ürünlerde yalnız en düşük kademe kalır.
def adet_sayisi(s):
    try:
        return int(str(s).split()[0].replace(".", ""))
    except (ValueError, IndexError):
        return 0

satirlar, matris = [], {}
for row in src.iter_rows(min_row=2, values_only=True):
    urun, tip, secenek, birim = row[1], row[2], row[4], row[5]
    slug = ESLE.get(urun)
    if slug and slug in fiyatli:
        continue                                   # sitede var → gereksiz
    if tip == "matris":
        k = (urun, secenek)
        n = adet_sayisi(birim)
        if k not in matris or n < matris[k][0]:
            matris[k] = (n, list(row))
    else:
        satirlar.append(list(row))
satirlar.extend(v[1] for v in matris.values())
satirlar.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[4])))

wb = openpyxl.Workbook()
kl = wb.active
kl.title = "ÖNCE BUNU OKU"
kl.column_dimensions["A"].width = 20
kl.column_dimensions["B"].width = 108
NOT = [
    ("NE DEĞİŞTİ", "Bu dosya sadeleştirildi. 658 satırlık ilk sürümde artık gereksiz olan 457 satır vardı — o ürünlerin fiyatı Viniltürk listesinden siteye girildi. Burada YALNIZ hâlâ eksik olanlar var."),
    ("", ""),
    ("NE YAPACAKSIN", "Sarı MALİYET sütununu doldur. SATIŞ isteğe bağlı — boş bırakırsan kategori kâr marjından hesaplanır."),
    ("", ""),
    ("ADET KADEMELERİ", "Matris ürünlerde yalnız EN DÜŞÜK kademe bırakıldı. Üst kademeleri senin kendi mevcut fiyatlarından çıkardığım katsayılarla türetirim: kartvizit ×1,93 (2 kat adet) / ×2,90 / ×4,79 / ×9,67 · broşür ×1,53 / ×2,88 / ×4,63 · çanta ×1,43. Katsayı sende farklıysa yaz, onu kullanırım."),
    ("", ""),
    ("BİRİM", "'1 adet (birim)' yazan satırlarda TEK ADETLİK fiyat yaz. Adet indirimi sistem tarafından otomatik uygulanır (10 adet %8, 25 adet %15, 50 adet %22, 100 adet %28, 250 adet %35). Adet adet fiyat YAZMA, yoksa indirim iki kez işler."),
    ("", ""),
    ("FARK SÜTUNU", "BİRİM hücresinde 'fark' yazan satırlara ANA fiyatı değil yalnızca FARKI yaz. Fark yoksa boş bırak, sıfır sayılır."),
    ("", ""),
    ("SATIR SİLME", "Satmadığın bir kalemin maliyetini boş bırak; silme. Boş bırakılan kombinasyon sitede hiç gösterilmez."),
    ("", ""),
    ("SADECE RAKAM", "1250 yaz — '1.250 TL' yazma."),
    ("", ""),
    ("EKSTRALAR", "Bu sayfa kaldırıldı. CNC kesim, laminasyon, iç mekan baskısı, kolon dikiş, dikiş+kopça ve germe Viniltürk fiyatlarıyla siteye zaten girildi."),
]
for a, b in NOT:
    kl.append([a, b])
for row in kl.iter_rows():
    if row[0].value:
        row[0].font = Font(bold=True)
    row[1].alignment = Alignment(wrap_text=True, vertical="top")

ws = wb.create_sheet("DOLDUR")
ws.append(basliklar)
for s in satirlar:
    ws.append(s)
for col, w in zip("ABCDEFGHI", [22, 30, 11, 24, 46, 16, 12, 12, 52]):
    ws.column_dimensions[col].width = w
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4B3AA0")
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{ws.max_row}"
sari = PatternFill("solid", fgColor="FFF9E0")
ince = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="hair"), bottom=Side(style="hair"))
for i in range(2, ws.max_row + 1):
    for c in (7, 8):
        cell = ws.cell(row=i, column=c)
        cell.fill = sari
        cell.border = ince
        cell.number_format = "#,##0.00"
    ws.cell(row=i, column=9).font = Font(size=9, color="777777")

wb.save(HEDEF)
print(f"{ws.max_row - 1} satır → {HEDEF}")
