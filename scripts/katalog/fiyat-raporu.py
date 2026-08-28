# -*- coding: utf-8 -*-
"""
TAM FİYAT RAPORU (2026-08-28) — Hasan talebi.

Sitedeki her ürünün her varyantı için maliyet ve satış fiyatını tek Excel'de toplar.

m² ÜRÜNLERİNDE SATIŞ SÜTUNU BOŞTUR — motor onu maliyetten türetir
(satış = maliyet × kur × marj × KDV). Rapor bu türetilmiş değeri "hesaplanan satış"
olarak ayrı sütunda gösterir ki tablo elle girilen ile hesaplananı karıştırmasın.

Kaynak: canlı veritabanı dışa aktarımı (§ ayraçlı).
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
KAYNAK = r"C:\Users\ADMINI~1\AppData\Local\Temp\fiyatlar.txt"
HEDEF = r"C:\Users\Administrator\Downloads\markala-fiyat-raporu.xlsx"
KUR, MARJ, KDV = 49.0, 1.2, 0.2

ALAN = ["katAd", "katSlug", "katAktif", "urunAd", "urunSlug", "urunAktif", "mod",
        "marj", "grup", "secenek", "birim", "etki", "adet", "maliyet", "satis"]

# İSG levhaları HARİÇ (Hasan 2026-08-28: "İSG ürünlerinin fiyatlarını çıkarmana
# gerek yok"). 828 ürün / ~19.850 satır; hepsi maliyetsiz ve ayrı yönetiliyor,
# rapora karıştıklarında gerçek katalog 20 bin satırın içinde kayboluyordu.
kayitlar = []
atlanan = 0
for satir in open(KAYNAK, encoding="utf-8"):
    p = satir.rstrip("\n").split("§")
    if len(p) != len(ALAN):
        continue
    k = dict(zip(ALAN, p))
    if k["katSlug"].startswith("is-guvenligi"):
        atlanan += 1
        continue
    kayitlar.append(k)
print(f"{len(kayitlar)} fiyat satırı okundu ({atlanan} İSG satırı atlandı)")


def sayi(s):
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


satirlar = []
for k in kayitlar:
    mal = sayi(k["maliyet"])
    sat = sayi(k["satis"])
    m2 = k["mod"] == "area"
    dolar = m2 and k["birim"] != "tl"
    # m² üründe satış sütunu 0'dır; gerçek satış motordan türetilir.
    hesap = None
    if m2 and mal:
        tl = mal * KUR if dolar else mal
        hesap = round(tl * MARJ * (1 + KDV), 2)
    gercek_satis = hesap if m2 else sat
    kar = None
    if mal and gercek_satis and mal > 0:
        temel = (mal * KUR if dolar else mal)
        kar = round(gercek_satis / (1 + KDV) / temel, 2) if temel else None
    satirlar.append([
        k["katAd"], "Evet" if k["katAktif"] == "t" else "Hayır",
        k["urunAd"], k["urunSlug"], "Evet" if k["urunAktif"] == "t" else "Hayır",
        {"area": "m²", "additive": "adet"}.get(k["mod"], k["mod"]),
        k["grup"], k["secenek"], k["adet"],
        mal, "USD" if dolar else "TRY",
        sat if not m2 else None,
        hesap, kar,
        float(k["marj"]) if k["marj"] else None,
    ])

BASLIK = ["Kategori", "Kat. Aktif", "Ürün", "Slug", "Ürün Aktif", "Fiyat Tipi",
          "Seçenek Grubu", "Seçenek", "Adet/Birim", "MALİYET", "Para Birimi",
          "SATIŞ (girilen)", "SATIŞ (hesaplanan m²)", "Gerçekleşen Marj", "Ürün Marjı"]

wb = openpyxl.Workbook()

# ── ÖZET ────────────────────────────────────────────────────────────────────
oz = wb.active
oz.title = "ÖZET"
oz.column_dimensions["A"].width = 46
oz.column_dimensions["B"].width = 14
oz.column_dimensions["C"].width = 60
maliyetli = sum(1 for s in satirlar if s[9])
maliyetsiz = len(satirlar) - maliyetli
urunler = {s[3] for s in satirlar}
ozet = [
    ("Rapor tarihi", "28.08.2026", "canlı veritabanından"),
    ("", "", ""),
    ("Toplam fiyat satırı", len(satirlar), ""),
    ("  maliyeti girilmiş", maliyetli, f"%{maliyetli/len(satirlar)*100:.1f}"),
    ("  MALİYETİ BOŞ", maliyetsiz, f"%{maliyetsiz/len(satirlar)*100:.1f} — kâr raporunda kârsız görünür"),
    ("", "", ""),
    ("Toplam ürün", len(urunler), ""),
    ("Kapsam", "İSG hariç", "İş güvenliği levhaları rapora dahil edilmedi (ayrı yönetiliyor)"),
    ("", "", ""),
    ("m² ürünlerinde satış", "hesaplanır", f"maliyet × kur({KUR:.0f}) × marj({MARJ}) × KDV(%{KDV*100:.0f})"),
    ("Dolar cinsinden maliyet", sum(1 for s in satirlar if s[10] == "USD"), "kur değişince fiyat kendiliğinden güncellenir"),
    ("", "", ""),
    ("SÜTUNLAR", "", ""),
    ("MALİYET", "", "sana mal oluş bedeli, KDV hariç"),
    ("SATIŞ (girilen)", "", "adet bazlı ürünlerde elle girilen satış fiyatı (KDV dahil)"),
    ("SATIŞ (hesaplanan m²)", "", "m² ürünlerinde motorun türettiği satış — elle girilmez"),
    ("Gerçekleşen Marj", "", "satış(KDV hariç) ÷ maliyet. 1,80 = %80 kâr"),
]
for r in ozet:
    oz.append(list(r))
for i in (1, 13):
    oz.cell(row=i, column=1).font = Font(bold=True)
for row in oz.iter_rows():
    if row[0].value and str(row[0].value).isupper():
        row[0].font = Font(bold=True)
    row[2].alignment = Alignment(wrap_text=True, vertical="top")
oz.cell(row=5, column=1).font = Font(bold=True, color="A32620")
oz.cell(row=5, column=3).font = Font(color="A32620")

# ── DETAY ───────────────────────────────────────────────────────────────────
ws = wb.create_sheet("TÜM FİYATLAR")
ws.append(BASLIK)
for s in satirlar:
    ws.append(s)
for col, w in zip("ABCDEFGHIJKLMNO", [24, 10, 34, 26, 11, 10, 22, 40, 13, 12, 11, 15, 20, 15, 11]):
    ws.column_dimensions[col].width = w
mor = PatternFill("solid", fgColor="4B3AA0")
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = mor
    c.alignment = Alignment(wrap_text=True, vertical="center")
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(BASLIK))}{ws.max_row}"
kirmizi = PatternFill("solid", fgColor="FBE9E7")
ince = Border(left=Side(style="hair"), right=Side(style="hair"))
for i in range(2, ws.max_row + 1):
    for c in (10, 12, 13):
        ws.cell(row=i, column=c).number_format = "#,##0.00"
        ws.cell(row=i, column=c).border = ince
    ws.cell(row=i, column=14).number_format = "0.00"
    if ws.cell(row=i, column=10).value in (None, ""):
        for c in range(1, len(BASLIK) + 1):
            ws.cell(row=i, column=c).fill = kirmizi   # maliyeti boş satır

wb.save(HEDEF)
print(f"{ws.max_row - 1} satır → {HEDEF}")
print(f"  maliyeti girilmiş: {maliyetli} · MALİYETİ BOŞ: {maliyetsiz} (kırmızı boyalı)")
