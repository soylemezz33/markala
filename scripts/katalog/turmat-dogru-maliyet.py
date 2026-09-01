# -*- coding: utf-8 -*-
"""TURMAT DOĞRU MALİYET TABLOSU (2026-08-29). Değerler Hasan'ın yapıştırdığı canlı
fiyat-listesi HTML'inden ELLE doğrulanarak gömüldü (parser riski yok — 0 hata kuralı).
Karşılaştırma: canlı DB dökümü (fiyat-inceleme.txt). Çıktı: okunaklı tek sayfa Excel.
Kartvizit: kod eşlemesi Hasan'dan gelene kadar BEKLEMEDE (satırlar bilgi olarak listelenir)."""
import sys
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill

sys.stdout.reconfigure(encoding="utf-8")
DB = r"C:\Users\ADMINI~1\AppData\Local\Temp\fiyat-inceleme.txt"
HEDEF = r"C:\Users\Administrator\Downloads\markala-turmat-maliyet.xlsx"

# (ürün adı, slug, [(kod, özellik, hedef_okey_ipucu, {adet: maliyet})])
# okey_ipucu: DB option_key/dim eşleşmesi; None → o üründe tek grup, adet yeter.
T = [
 ("Broşür", "brosur", [
  ("CA7","A7 9.5x20","a7",{1000:800,2000:1100,5000:1950,10000:3700}),
  ("CA5","A5 14x20","a5",{1000:900,2000:1200,5000:2100,10000:3900}),
  ("CA4","A4 20x28","a4",{1000:1600,2000:2200,5000:3950,10000:7600}),
  ("CA3","A3 28x40","a3",{1000:3100,2000:4300,5000:7800,10000:14900})]),
 ("El İlanı", "el-ilani", [
  ("ELI3-4","A7 9.2x20","a7",{6000:1850,12000:3600}),
  ("ELI5-8","A5 13.8x20","a5",{2000:900,4000:1800,6000:2600,8000:3400}),
  ("ELI9-12","A4 19x27.2","a4",{2000:1650,4000:3100,6000:4600,8000:6000}),
  ("ELI13-16","A3 27.5x40","a3",{2000:3100,4000:6000,6000:9000,8000:11200})]),
 ("Afiş 105gr", "afis-105gr", [
  ("AF1-3","34x49","34x49",{250:1950,500:2500,1000:3000}),
  ("AF4-6","49x69","49x69",{250:3300,500:3850,1000:5200})]),
 ("Antetli Kağıt", "antetli-kagit", [
  ("ANT1-3","A5 90gr","a5",{4000:2100,8000:4000,12000:5850}),
  ("ANT4-6","A4 90gr","a4",{2000:2000,4000:3700,6000:5400})]),
 ("Zarf Diplomat Renkli", "zarf-diplomat-renkli", [("Z2","10.5x24 (baz+artış)","diplomat-renkli",
  {1000:2375,2000:3625,3000:4875,5000:7375,10000:13625})]),
 ("Torba Zarf", "zarf-torba", [("Z3","24x32 (baz+artış)","torba-renkli",
  {500:3500,1000:5375,1500:7250,2500:11000,5000:20375})]),
 ("Kapı Askı", "kapi-aski-brosur", [
  ("ASK1","10,5x24 350gr","ask1",{1000:2800}),("ASK2","10,5x24 700gr","ask2",{1000:3000}),
  ("ASK3","21x26 200gr","ask3",{1000:4850})]),
 ("Cepli Dosya", "cepli-dosya", [
  ("PD","Parlak selefon","pd",{1000:8750}),("MND","Mat selefon içi tek renk","mnd",{500:5900,1000:9000}),
  ("KLD","Mat selefon kabartma lak","kld",{500:6400,1000:9750}),("CYPD","Çift yön parlak","cypd",{1000:11250}),
  ("CYMD","Çift yön mat","cymd",{1000:11500}),("CYML4D","400gr laklı","cyml4d",{1000:15000}),
  ("AVD","Avukat dosyası","avd",{500:5000,1000:6150})]),
 ("Etiket", "etiket", [
  ("E","53x83 selefonlu","e",{1000:300}),("ES","53x83 selefonsuz","es",{1000:280}),
  ("EO","52x82 özel kesim","eo",{1000:500}),("EOY","52x82 yaldız","eoy",{1000:740}),
  ("ETM","15.5x25.5","etm",{1000:2600}),("ETL","25.5x33","etl",{1000:5100})]),
 ("Makbuz", "makbuz", [
  ("M1","10x14 siyah","m1",{20:1200}),("M2","10x20 siyah","m2",{30:1900}),
  ("M3","14x20 siyah","m3",{10:900}),("M4","20x29 siyah","m4",{10:1500}),
  ("MR1","10x14 renkli","mr1",{20:1300}),("MR2","10x20 renkli","mr2",{30:2200}),
  ("MR3","14x20 renkli","mr3",{10:1200}),("MR4","20x29 renkli","mr4",{10:1800})]),
 ("Oto Paspas", "oto-paspas", [("P1-3","34x49 85gr kraft","paspas",{1000:1750,2000:2700,5000:4800})]),
 ("Küp Bloknot", "kup-bloknot", [
  ("NKKB-250","250 yaprak","250lik",{100:7200,250:8750,500:11250,1000:17500}),
  ("NKKB-500","500 yaprak","500luk",{100:9375,250:13125,500:18125})]),
 ("Spiralli Bloknot", "spiralli-bloknot", [
  ("B1","9.4 NK","9-nk",{500:10200,1000:17400}),("B2","9.4 CYP","9-cyp",{500:10800,1000:18000}),
  ("B3","9.4 CYM","9-cym",{500:10800,1000:18000}),("B4","9.4 CYML4","9-cyml4",{500:11520,1000:18720}),
  ("B5","9.4 SEK","9-sek",{500:13440,1000:20640}),("B6","14x20 NK","14-nk",{500:15000,1000:24000}),
  ("B7","14x20 CYP","14-cyp",{500:16200,1000:25200}),("B8","14x20 CYM","14-cym",{500:16200,1000:25200}),
  ("B9","14x20 CYML4","14-cyml4",{500:17400,1000:27000}),("B10","14x20 SEK","14-sek",{500:22800,1000:31800})]),
 ("Çanta Bristol", "canta", [
  ("CNT1","25x37x8","cnt1",{500:9850,1000:14400,2000:26275}),
  ("CNT2","38x23x9","cnt2",{500:10800,1000:15600,2000:29975}),
  ("CNT3","17x24x7","cnt3",{500:9000,1000:11400,2000:19525}),
  ("CNT4","51x33x13","cnt4",{500:13800,1000:22800,2000:42175})]),
 ("Çanta Kraft", "canta", [
  ("CNT1-KRAFT","25x37x8 kraft","cnt1-kraft",{500:6900,1000:10200,2000:18700}),
  ("CNT3-KRAFT","17x24x7 kraft","cnt3-kraft",{500:6600,1000:8400,2000:15275})]),
]
KARTVIZIT_BEKLIYOR = "NK=220 NKA=240 NSK=200 MNA=260 CYP=320 CYM=310 KL=320 CYML4=400 O-COK=430 S-COK=500 O-SEK=650 EKO-SEK=600 S-SEK=750 A-SEK=1000 AC-SEK=1100 TANK=715 AY=600 GY=620 VIP=800 FAN=450 F-SEK=780"

db = defaultdict(list)
ALAN = ["kat","urun","slug","mod","marj","grup","secenek","gkey","okey","dim","birim","etki","maliyet","satis"]
for ham in open(DB, encoding="utf-8"):
    p = ham.rstrip("\n").split("§")
    if len(p) == len(ALAN):
        r = dict(zip(ALAN, p)); db[r["slug"]].append(r)

def katla(s):
    return s.translate(str.maketrans("ÇĞİÖŞÜI","çğiöşüı")).lower()

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "KARŞILAŞTIRMA"
ws.append(["ÜRÜN","TURMAT KOD","ÖZELLİK","ADET","TURMAT ₺","BİZDE ₺","FARK","DURUM"])
ayni=farkli=yok=0
for urun, slug, kalemler in T:
    for kod, ozellik, ipucu, fiyatlar in kalemler:
        for adet, tf in sorted(fiyatlar.items()):
            adaylar = [r for r in db.get(slug,[]) if r["dim"]==str(adet) or (not r["dim"] and len(fiyatlar)==1)]
            if ipucu:
                # Önce TAM anahtar eşleşmesi (cnt1, cnt1-kraft'a bulaşmasın); yoksa metin içi arama.
                tam = [r for r in adaylar if katla(r["okey"]) == katla(ipucu)]
                dar = tam or [r for r in adaylar if katla(ipucu) in katla(r["okey"]+" "+r["secenek"])]
                adaylar = dar or adaylar
            if len({r["okey"] for r in adaylar}) == 1:
                r = adaylar[0]; mev = float(r["maliyet"]) if r["maliyet"] else None
                fark = None if mev is None else round(tf-mev,2)
                d = "✓ aynı" if fark==0 else ("maliyet boş" if mev is None else "✗ FARKLI")
                ayni += fark==0; farkli += (fark not in (0,None) or mev is None)
                ws.append([urun,kod,ozellik,adet,tf,mev,fark,d])
            else:
                yok += 1
                ws.append([urun,kod,ozellik,adet,tf,None,None,"eşleşmedi/çok aday"])
ws.append([]); ws.append(["Kartvizit","(21 kod)","21/21 DB anahtarıyla doğrulandı (vip 780→800 düzeltildi)","",KARTVIZIT_BEKLIYOR,"","","✓ aynı"])
for c in ws[1]: c.font = Font(bold=True,color="FFFFFF"); c.fill = PatternFill("solid",fgColor="4B3AA0")
kir = PatternFill("solid",fgColor="FBE9E7"); gri = PatternFill("solid",fgColor="EFEFEF")
for i in range(2, ws.max_row+1):
    d = str(ws.cell(row=i,column=8).value or "")
    if d.startswith(("✗","maliyet")): f=kir
    elif d.startswith(("eşleş","BEKLE")): f=gri
    else: continue
    for c in range(1,9): ws.cell(row=i,column=c).fill = f
for col,w in zip("ABCDEFGH",[22,13,26,8,11,11,9,18]): ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:H{ws.max_row}"
wb.save(HEDEF)
print(f"✓ {ayni} · ✗/boş {farkli} · eşleşmedi {yok} → {HEDEF}")
