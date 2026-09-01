# -*- coding: utf-8 -*-
"""
FİYAT GÖZDEN GEÇİRME DOSYASI (2026-08-29) — Hasan talebi:
"tüm ürünlerin maliyet ve satış fiyatının olduğu exceli iletebilir misin
doldurup sana verelim" — yalnız eksikler değil, HER ŞEY.

markala-fiyat-girisi.xlsx'ten farkı: mevcut değerler DOLU gelir, Hasan
beğenmediği hücreyi DEĞİŞTİRİR. Geri gelince fark çıkarılır, önizleme
onaylanınca siteye yazılır. İSG hariç (Hasan 2026-08-28), pasif ürünler hariç.

Import anahtarları (SLUG/GRUP-K/SEÇENEK-K/KADEME sütunları) sağ blokta —
bunlara dokunulmaz; satır silme/ekleme yok.
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
# Git Bash /tmp → Windows gerçek konumu (cygpath -w ile doğrulandı).
KAYNAK = r"C:\Users\ADMINI~1\AppData\Local\Temp\fiyat-inceleme.txt"
KAYNAK2 = r"C:\Users\ADMINI~1\AppData\Local\Temp\fiyat-tekli.txt"
HEDEF = r"C:\Users\Administrator\Downloads\markala-fiyat-inceleme.xlsx"
KUR, MARJ_VARSAYILAN, KDV = 49.0, 1.2, 0.2

ALAN = ["kat", "urun", "slug", "mod", "marj", "grup", "secenek",
        "gkey", "okey", "dim", "birim", "etki", "maliyet", "satis"]

satirlar = []
for yol in (KAYNAK, KAYNAK2):
    try:
        f = open(yol, encoding="utf-8")
    except OSError:
        f = open(yol.replace("/tmp/", "C:/cygwin64/tmp/"), encoding="utf-8")
    for ham in f:
        p = ham.rstrip("\n").split("§")
        if len(p) != len(ALAN):
            continue
        satirlar.append(dict(zip(ALAN, p)))
print(f"{len(satirlar)} fiyat satırı okundu")


def sayi(s):
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


wb = openpyxl.Workbook()

# ── OKU ─────────────────────────────────────────────────────────────────────
oz = wb.active
oz.title = "ÖNCE BUNU OKU"
oz.column_dimensions["A"].width = 22
oz.column_dimensions["B"].width = 100
kurallar = [
    ("NE BU", "Sistemdeki TÜM aktif ürünlerin güncel maliyet + satış fiyatları (İSG levhaları hariç). "
              "Beğenmediğin MALİYET/SATIŞ hücresini değiştir, dosyayı bana geri ver. "
              "Farkları çıkarıp önce ONAY TABLOSU göstereceğim — onaylamadan siteye hiçbir şey yazılmaz."),
    ("", ""),
    ("DEĞİŞTİREBİLECEĞİN", "Yalnız iki sütun: MALİYET ve SATIŞ. Başka hücreye dokunma."),
    ("m² ÜRÜNLERİ", "Fiyat Tipi 'm²' olan satırlarda SATIŞ sütunu 'hesaplanır' yazar ve DEĞİŞTİRİLMEZ — "
                    "motor satışı maliyetten türetir (maliyet × kur × marj × KDV). Bu ürünlerde yalnız MALİYETİ düzelt. "
                    "REFERANS sütunu bugünkü kurla (49₺) 1 m²/1 adet satışının kaça denk geldiğini gösterir."),
    ("PARA BİRİMİ", "PARA sütunu USD diyorsa maliyeti DOLAR yaz (kur değişince fiyat kendini günceller), "
                    "TRY diyorsa TL yaz. Birimi değiştirme."),
    ("SATIŞ KDV", "SATIŞ fiyatları KDV DAHİL, müşteri ekranda ne görüyorsa o. MALİYET her zaman KDV hariç."),
    ("BOŞ HÜCRE", "Boş bıraktığın hücre DEĞİŞTİRİLMEZ (mevcut değer korunur). Bir değeri silmek istiyorsan 0 yaz."),
    ("SATIR SİLME", "Satır silme/ekleme YOK — yeni ürün/varyant işi markala-fiyat-girisi.xlsx'te."),
    ("SAĞDAKİ GRİ BLOK", "SLUG'dan sonraki gri sütunlar sistemin eşleştirme anahtarları — DOKUNMA."),
    ("", ""),
    ("İPUCU", "Üst satırdaki filtre oklarıyla kategori/ürün seçip bölüm bölüm ilerleyebilirsin. "
              "MARJ sütunu o satırın bugünkü gerçekleşen kâr çarpanı (1,80 = %80) — düşük görünenler ilk bakılacak yerler."),
]
for r in kurallar:
    oz.append(list(r))
for row in oz.iter_rows():
    row[0].font = Font(bold=True)
    row[1].alignment = Alignment(wrap_text=True, vertical="top")

# ── FİYATLAR ────────────────────────────────────────────────────────────────
ws = wb.create_sheet("FİYATLAR")
BASLIK = ["KATEGORİ", "ÜRÜN", "FİYAT TİPİ", "SEÇENEK GRUBU", "SEÇENEK", "KADEME/ADET",
          "PARA", "MALİYET", "SATIŞ", "ÖNERİ SATIŞ", "REFERANS (m² satış)", "MARJ",
          "SLUG", "GRUP-K", "SEÇENEK-K", "KADEME-K"]
# Hedef marj: ürün marjı varsa o, yoksa matbaa hedefi 1.70 (Hasan aralığı 1.65-1.80).
HEDEF_MARJ = 1.7
ws.append(BASLIK)

for k in satirlar:
    mal = sayi(k["maliyet"])
    sat = sayi(k["satis"])
    m2 = k["mod"] == "area"
    dolar = m2 and k["birim"] != "tl"
    ref = None
    if m2 and mal:
        tl = mal * KUR if dolar else mal
        # Ürün marjı varsa onu, yoksa m² varsayılanını kullan (bilgi amaçlı referans).
        m = sayi(k["marj"]) or MARJ_VARSAYILAN
        ref = round(tl * m * (1 + KDV), 2)
    marj_g = None
    if mal and mal > 0:
        temel = mal * KUR if dolar else mal
        kaynak_satis = ref if m2 else sat
        if kaynak_satis and temel:
            marj_g = round(kaynak_satis / (1 + KDV) / temel, 2)
    # ÖNERİ SATIŞ (2026-08-29, Hasan: "maliyeti görüp fiyatları baştan belirlemek"):
    # maliyet(KDV hariç TL) × hedef marj × KDV, 10₺'ye yukarı yuvarlı. m²'de boş (motor türetir).
    oneri = None
    if not m2 and mal and mal > 0:
        m = sayi(k["marj"]) or HEDEF_MARJ
        oneri = -(-int(mal * m * (1 + KDV)) // 10) * 10
    ws.append([
        k["kat"], k["urun"],
        {"area": "m²", "additive": "adet", "matris": "matris"}.get(k["mod"], k["mod"]),
        k["grup"] or "—", k["secenek"] or "—", k["dim"] or "",
        "USD" if dolar else "TRY",
        mal,
        "hesaplanır" if m2 else sat,
        oneri, ref, marj_g,
        k["slug"], k["gkey"], k["okey"], k["dim"],
    ])

genislikler = [20, 34, 9, 20, 30, 12, 7, 11, 11, 12, 15, 7, 26, 12, 18, 10]
for i, wdt in enumerate(genislikler, start=1):
    ws.column_dimensions[get_column_letter(i)].width = wdt

mor = PatternFill("solid", fgColor="4B3AA0")
sari = PatternFill("solid", fgColor="FFF3CD")
gri = PatternFill("solid", fgColor="EFEFEF")
ince = Border(left=Side(style="hair"), right=Side(style="hair"))
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = mor
    c.alignment = Alignment(wrap_text=True, vertical="center")
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(BASLIK))}{ws.max_row}"

for i in range(2, ws.max_row + 1):
    for col in (8, 9):  # MALİYET, SATIŞ — düzenlenebilir alanlar sarı
        h = ws.cell(row=i, column=col)
        if h.value != "hesaplanır":
            h.fill = sari
        h.number_format = "#,##0.00"
        h.border = ince
    ws.cell(row=i, column=10).number_format = "#,##0"; ws.cell(row=i, column=11).number_format = "#,##0.00"
    ws.cell(row=i, column=12).number_format = "0.00"
    for col in (13, 14, 15, 16):  # anahtar blok — gri, dokunulmaz
        ws.cell(row=i, column=col).fill = gri
        ws.cell(row=i, column=col).font = Font(color="999999", size=9)

wb.save(HEDEF)
print(f"{ws.max_row - 1} satır → {HEDEF}")
