# -*- coding: utf-8 -*-
"""
FİYAT GİRİŞ ŞABLONU ÜRETİCİ (2026-08-28, Hasan talebi).

Hasan'ın kategori Excel'ini okur ve siteye doğrudan import edilebilecek düz bir fiyat
tablosu üretir. Hasan yalnız MALİYET (ve isterse SATIŞ) sütununu doldurur.

ÜÇ FİYAT TİPİ VAR — çünkü sitedeki motor (apps/api/src/orders/pricing.ts) üç türlü
hesaplıyor. Yanlış tipte satır girilirse fiyat yanlış çıkar:

  matris     her (seçenek × adet) çifti için AYRI toplam fiyat. Hacim indirimi
             UYGULANMAZ; kademe fiyatını zaten sen belirlemişsindir.
  toplamali  yalnız BİRİM fiyat. Seçili "priced" grupların fiyatları TOPLANIR, sonra
             adetle çarpılır ve hacim indirimi OTOMATİK iner (10→%8 … 250→%35).
  m2         1 m² fiyatı. Kenar işlemleri çevre metresi, aksesuarlar adet başına.

Kullanım:
  python scripts/katalog/sablon-uret.py <kaynak.xlsx> <hedef.xlsx>
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

KAYNAK = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\Administrator\Downloads\dijital baskı kategorisi (1).xlsx"
HEDEF = sys.argv[2] if len(sys.argv) > 2 else r"C:\Users\Administrator\Downloads\markala-fiyat-girisi.xlsx"

KADEME = {
    "kartvizit": [1000, 2000, 3000, 5000, 10000],
    "matbaa":    [1000, 2000, 5000, 10000],
    "canta":     [500, 1000, 2000],
    "bloknot":   [100, 250, 500, 1000],
}

# ürün adı -> (fiyat tipi, adet kademesi). Listede olmayan her ürün "toplamali" sayılır.
TIP = {
    "Avrupa Vinil Baskı": ("m2", None), "Çin (Lamine) Vinil Baskı": ("m2", None),
    "Mesh Delikli Vinil Baskı": ("m2", None), "One Way Vision Baskı": ("m2", None),
    "Folyo Baskı": ("m2", None), "Baskes Folyo": ("m2", None), "Duvar Kağıdı": ("m2", None),
    "Dekota\\Foreks Baskı": ("m2", None), "Pleksi Baskı": ("m2", None),
    "Kompozit Baskı": ("m2", None), "Kanvas Tablo Baskı": ("m2", None), "UV DTF Baskı": ("m2", None),

    "Sticker Baskı": ("matris", "matbaa"), "Amerikan Servis": ("matris", "matbaa"),
    "Islak Mendil": ("matris", "matbaa"), "Baskılı Stick Şeker": ("matris", "matbaa"),
    "Baskılı Sachet Tuz": ("matris", "matbaa"), "Bardak Altlığı": ("matris", "matbaa"),
    "Masa Kartı - Tent Card": ("matris", "matbaa"), "Adisyon Fişi": ("matris", "matbaa"),
    "Kağıt Oto Paspas": ("matris", "matbaa"), "Antetli Kağıt": ("matris", "matbaa"),
    "Diplomat Zarf": ("matris", "matbaa"), "Cepli Dosya": ("matris", "matbaa"),
    "Makbuz": ("matris", "matbaa"), "El İlanı": ("matris", "matbaa"),
    "Kapı Askısı El İlanı": ("matris", "matbaa"), "Broşür": ("matris", "matbaa"),
    "Spiralli Bloknot/Defter": ("matris", "bloknot"), "Masa Takvimi": ("matris", "bloknot"),
    "Küp Bloknot": ("matris", "bloknot"),
    "Amerikan Bristol Karton Çanta": ("matris", "canta"), "Kraft Karton Çanta": ("matris", "canta"),
    "Standart Kartvizit": ("matris", "kartvizit"), "Sıvama Kartvizit": ("matris", "kartvizit"),
    "Kabartma Laklı Kartvizit": ("matris", "kartvizit"), "Şeffaf Kartvizit": ("matris", "kartvizit"),
}

# ── Kaynağı ayrıştır ────────────────────────────────────────────────────────
wb_in = openpyxl.load_workbook(KAYNAK, data_only=True)
urunler = []
for ws in wb_in.worksheets:
    cur = None
    onceki_tek_hucre = False
    for row in ws.iter_rows(values_only=True):
        c = ["" if x is None else str(x).strip() for x in row]
        if not any(c):
            onceki_tek_hucre = False                     # boş satır → blok bitti
            continue
        if c[0] and not any(c[1:]):                      # yalnız ilk hücre dolu
            # ÜST ÜSTE iki tek-hücreli satır: ikincisi YENİ ÜRÜN DEĞİL, DEĞERİ BOŞ
            # BIRAKILMIŞ bir seçenek grubudur. Kaşelerde böyle: "Standart Kaşe" (ürün)
            # → "Marka" (değerleri yazılmamış grup) → "Ebat" / "Renk" (ürünün grupları).
            # Bu ayrım olmadan Ebat ve Renk, ürün sanılan "Marka"ya bağlanıyordu.
            if onceki_tek_hucre and cur is not None:
                cur["bos_gruplar"].append(c[0])
            else:
                cur = {"sayfa": ws.title, "urun": c[0], "gruplar": [], "bos_gruplar": []}
                urunler.append(cur)
            onceki_tek_hucre = True
            continue
        onceki_tek_hucre = False
        if cur is not None and c[0]:
            cur["gruplar"].append({"ad": c[0], "secenekler": [v for v in c[1:] if v]})

# Seçenek grubuna göre BİRİM. m² ürünlerde her grup metrekareyle fiyatlanmaz:
# dikiş/germe kenar boyunca (çevre metresi), kopça/direk/duba adet başınadır.
# Motor bunların hepsini destekliyor (rules.effect: perM2 / perPerimeter / perPiece).
def katla(s):
    """Türkçe katlama. 'İ'.lower() 'i̇' (i + birleşik nokta) verir; düz `in` araması
    bu yüzden 'Ek İşlemler' içinde 'ek işlem' bulamaz. Önce harfleri sadeleştiriyoruz."""
    esle = str.maketrans("İIıŞşĞğÜüÖöÇç", "iiisSgGuUoOcc")
    return s.translate(esle).lower()


def birim_of(tip, grup_ad):
    if tip != "m2":
        return "1 adet (birim)"
    g = katla(grup_ad)
    if "ek islem" in g:
        return "1 metretül (çevre)"
    if "cnc" in g:
        return "1 metretül (kesim yolu)"
    if "baski turu" in g:
        return "1 m²"
    return "1 m²"


# EKSTRALAR tüm ürünlerde ORTAK tek listedir — her üründe tekrar tekrar sorulmaz.
# Birimleri FARKLI: dikiş/kopça kenar boyunca (çevre), CNC kesim levha alanı üzerinden.
EKSTRA_GRUPLARI = {"Ek İşlemler", "CNC Kesim"}
EKSTRA_BIRIM = {
    "Sadece Dikiş":  ("metretül (çevre)", "Kenar boyunca dikiş"),
    "Sadece Kopça":  ("metretül (çevre)", "Kenara ~50 cm arayla kopça"),
    "Dikiş + Kopça": ("metretül (çevre)", "Dikiş ve kopça birlikte"),
    "Germe":         ("metretül (çevre)", "Kenar germe / gergi"),
    "Kolon + Dikiş": ("metretül (çevre)", "Kolon (kayış) + dikiş"),
    "Evet":          ("m²", "CNC kesim — levha alanı üzerinden"),
}
ekstralar = {}   # secenek -> (grup, birim, aciklama)

# ── Satırları üret ──────────────────────────────────────────────────────────
satirlar = []
for u in urunler:
    tip, kademe_ad = TIP.get(u["urun"], ("toplamali", None))
    kademeler = KADEME.get(kademe_ad) if kademe_ad else None
    # Ekstraları ürünün satırlarından ÇIKAR, ortak listeye topla.
    for g in u["gruplar"]:
        if g["ad"] in EKSTRA_GRUPLARI:
            for s in g["secenekler"]:
                if s == "Hayır":
                    continue          # "CNC Kesim: Hayır" ücretsizdir, satır gerekmez
                birim, acik = EKSTRA_BIRIM.get(s, ("metretül (çevre)", ""))
                # "CNC Kesim: Evet" tek başına anlamsız bir etiket → grubun adını kullan.
                ad = g["ad"] if s == "Evet" else s
                ekstralar[ad] = (g["ad"], birim, acik)
    gruplar_ek_haric = [g for g in u["gruplar"] if g["ad"] not in EKSTRA_GRUPLARI]

    # Tek seçenekli gruplar fiyat taşımaz — ürünün sabit özelliğidir.
    fiyatli = [g for g in gruplar_ek_haric if len(g["secenekler"]) > 1]
    sabit = " · ".join(f'{g["ad"]}: {g["secenekler"][0]}'
                       for g in gruplar_ek_haric if len(g["secenekler"]) == 1)
    # Değerleri yazılmamış gruplar (ör. kaşede "Marka") — doldurulması gerektiği görünsün.
    if u.get("bos_gruplar"):
        eksik_not = "DEĞERLER EKSİK → " + ", ".join(u["bos_gruplar"])
        sabit = f"{sabit} · {eksik_not}" if sabit else eksik_not

    # m² ÜRÜNLERDE gruplar ÇARPILIR, toplanmaz: 5 mm siyah pleksi tek bir malzemedir,
    # "5 mm" + "siyah farkı" değil. Aynı şekilde malzeme × baskı türü tek satırdır.
    if tip == "m2" and len(fiyatli) > 1:
        from itertools import product as _carp
        adlar = " × ".join(g["ad"] for g in fiyatli)
        for kombin in _carp(*[g["secenekler"] for g in fiyatli]):
            satirlar.append([u["sayfa"], u["urun"], tip, adlar, " — ".join(kombin),
                             "1 m²", None, None, sabit])
        continue
    if not fiyatli:
        adet = f"{kademeler[0]} adet" if tip == "matris" else birim_of(tip, "")
        satirlar.append([u["sayfa"], u["urun"], tip, "—", "(tek varyant)", adet, None, None, sabit])
        continue
    # Birden çok seçenek grubu varsa ANA fiyat İLK gruba yazılır; sonrakiler yalnız FARK.
    # (Motor priced grupların satırlarını topluyor — pricing.ts "Birim = Σ priced gruplar".)
    for gi, g in enumerate(fiyatli):
        fark = " · fark" if gi > 0 else ""
        for s in g["secenekler"]:
            if tip == "matris" and kademeler:
                for a in kademeler:
                    satirlar.append([u["sayfa"], u["urun"], tip, g["ad"], s, f"{a} adet{fark}", None, None, sabit])
            else:
                satirlar.append([u["sayfa"], u["urun"], tip, g["ad"], s, birim_of(tip, g["ad"]) + fark, None, None, sabit])

# ── Çıktı ───────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

kl = wb.active
kl.title = "KILAVUZ"
kl.column_dimensions["A"].width = 16
kl.column_dimensions["B"].width = 112
KILAVUZ = [
    ("FİYAT TİPİ", "NE GİRECEKSİN"),
    ("matris", "Her (seçenek × adet) satırına O ADEDİN TOPLAM fiyatını yaz. Örn. 1000 adet kartvizit 350, 2000 adet 700. Hacim indirimi uygulanmaz — kademeyi zaten sen belirliyorsun."),
    ("toplamali", "Yalnız BİRİM fiyat (1 adetlik) yaz. Seçilen grupların fiyatları toplanır. Hacim indirimi OTOMATİKTİR: 10 adet %8, 25 adet %15, 50 adet %22, 100 adet %28, 250 adet %35. Adet başına satır YAZMA, yoksa indirim iki kez uygulanır."),
    ("m2", "1 metrekarenin fiyatını yaz. Kenar işlemleri (dikiş, kopça) çevre metresi üzerinden, aksesuarlar (direk, duba) adet başına ayrı satırdır — BİRİM sütununda yazıyor."),
    ("", ""),
    ("ANA FİYAT / FARK", "Bir üründe birden çok seçenek grubu varsa ANA fiyatı İLK gruba yaz; sonraki gruplara yalnızca FARKI yaz. Fark yoksa boş bırak — sıfır sayılır. BİRİM sütununda 'fark' yazan satırlar bunlardır. Örnek: Polo tişörtün fiyatı 'Baskı Yönü'ne yazılır; 'Beden'de yalnız XXL ek ücretliyse o yazılır, S-M-L-XL boş kalır. Kartvizitte fiyat 'Baskı Yönü'ne (tek/çift yüz) yazılır, 'Tasarım Yönü' fiyata etki etmiyorsa boş kalır."),
    ("BİRİM'e dikkat", "m² ürünlerde her satır metrekare değildir. 'metretül (çevre)' yazan satır kenar boyunca fiyatlanır (dikiş, germe): 100x200 cm brandanın çevresi 6 metredir. 'metretül (kesim yolu)' CNC kesim içindir. Kopça, direk, duba gibi aksesuarları adet başına fiyatlayacaksan BİRİM hücresini '1 adet (aksesuar)' olarak düzelt."),
    ("", ""),
    ("SÜTUNLAR", ""),
    ("MALİYET", "ZORUNLU. Sana mal oluş bedeli, KDV HARİÇ. Kâr raporu bu sütundan hesaplanır; boş kalırsa o satır kârsız görünür."),
    ("SATIŞ", "İsteğe bağlı. Boş bırakırsan kategori kâr marjından otomatik hesaplanır (kartvizit 1.65 · broşür 1.70 · ambalaj 1.80). Elle yazarsan seninki geçerli olur."),
    ("", ""),
    ("KURALLAR", ""),
    ("1", "Satır SİLME, boş bırak. Bir kademeyi satmıyorsan MALİYET'i boş bırak — o kombinasyon sitede hiç görünmez. Silersen hangi kombinasyonu atladığın kaybolur."),
    ("2", "Adet kademesi ekleyebilirsin: satırı kopyala, BİRİM / ADET hücresini düzenle."),
    ("3", "Sadece rakam yaz. 'TL', binlik noktası, boşluk koyma. 1250 yaz — '1.250 TL' yazma."),
    ("4", "Tek seçenekli gruplar (ör. 'Baskı Yönü: Tek Yön') fiyat satırı almaz; SABİT ÖZELLİKLER sütununda ürün açıklaması olarak durur."),
    ("5", "Yeni ürün/seçenek eklemek istersen en alta yaz; SAYFA, ÜRÜN ve FİYAT TİPİ sütunlarını doldurman yeterli."),
    ("6", "Sütun başlıklarını ve sıralamasını DEĞİŞTİRME — import bunlara göre okuyor."),
]
for r in KILAVUZ:
    kl.append(list(r))
for i in (1, 6, 9, 10, 13):
    kl.cell(row=i, column=1).font = Font(bold=True)
    kl.cell(row=i, column=2).font = Font(bold=True)
for row in kl.iter_rows():
    if row[0].value:
        row[0].font = Font(bold=True)
    row[1].alignment = Alignment(wrap_text=True, vertical="top")

ek = wb.create_sheet("EKSTRALAR")
ek.append(["EK İŞLEM", "GRUP", "BİRİM", "MALİYET", "SATIŞ", "AÇIKLAMA"])
for s, (grup, birim, acik) in sorted(ekstralar.items(), key=lambda x: (x[1][0], x[0])):
    ek.append([s, grup, birim, None, None, acik])
for col, w in zip("ABCDEF", [22, 16, 20, 12, 12, 46]):
    ek.column_dimensions[col].width = w
for c in ek[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4B3AA0")
for i in range(2, ek.max_row + 1):
    for c in (4, 5):
        ek.cell(row=i, column=c).fill = PatternFill("solid", fgColor="FFF9E0")
        ek.cell(row=i, column=c).border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                                 top=Side(style="hair"), bottom=Side(style="hair"))
        ek.cell(row=i, column=c).number_format = "#,##0.00"
ek.freeze_panes = "A2"
ek["A" + str(ek.max_row + 2)] = "Bu liste TÜM ürünler için ORTAKtır — bir kez doldur, bu seçeneğin olduğu her üründe kullanılır."
ek["A" + str(ek.max_row)].font = Font(italic=True, size=9, color="777777")

ws = wb.create_sheet("FİYAT GİRİŞİ")
BASLIK = ["SAYFA", "ÜRÜN", "FİYAT TİPİ", "SEÇENEK GRUBU", "SEÇENEK",
          "BİRİM / ADET", "MALİYET", "SATIŞ", "SABİT ÖZELLİKLER"]
ws.append(BASLIK)
for r in satirlar:
    ws.append(r)

for col, w in zip("ABCDEFGHI", [22, 30, 11, 24, 46, 16, 12, 12, 52]):
    ws.column_dimensions[col].width = w
mor = PatternFill("solid", fgColor="4B3AA0")
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = mor
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{ws.max_row}"

RENK = {"m2": "E8E3F5", "matris": "E3EFE8", "toplamali": "F7EEDD"}
sari = PatternFill("solid", fgColor="FFF9E0")
ince = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="hair"), bottom=Side(style="hair"))
for i in range(2, ws.max_row + 1):
    tip = ws.cell(row=i, column=3).value
    if tip in RENK:
        ws.cell(row=i, column=3).fill = PatternFill("solid", fgColor=RENK[tip])
    for c in (7, 8):
        cell = ws.cell(row=i, column=c)
        cell.fill = sari
        cell.border = ince
        cell.number_format = "#,##0"
    ws.cell(row=i, column=9).font = Font(size=9, color="777777")

wb.save(HEDEF)

sayim = {}
for s in satirlar:
    sayim[s[2]] = sayim.get(s[2], 0) + 1
print(f"{len(satirlar)} fiyat satiri, {len(urunler)} urun -> {HEDEF}")
print("  tip dagilimi:", sayim)
