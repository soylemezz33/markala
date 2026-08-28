# -*- coding: utf-8 -*-
"""
DİJİTAL BASKI KATALOĞUNU TAMAMLA (2026-08-28).

Maliyetler Viniltürk dosyasından OKUNUR, elle yazılmaz — 20/20 doğrulama tuttuğu
için o dosya sitenin maliyet kaynağı (bkz. sohbet 28.08).

  A) Mevcut ürünlere eksik malzemeler (Folyo Çeşitleri + Vinil Branda UV türleri)
  B) 7 yeni ürün: Baskes Folyo · Duvar Kağıdı · Pleksi · Kompozit · Kanvas Tablo ·
     UV DTF · One Way Vision
  C) Ek işlemler: CNC Kesim, Laminasyon, İç Mekan Baskı, Kolon Dikiş, dikiş+kopça, germe

m² ÜRÜNLERDE SATIŞ FİYATI GİRİLMEZ. Motor `satış = maliyet × kur × marj × KDV`
hesaplıyor (pricing.ts computeAreaPrice), price sütunu okunmuyor → 0 yazılır.
Maliyet DOLAR cinsinden tutulur (rules.birim="dolar"); kur değişince fiyat
kendiliğinden güncellenir — mevcut 29 m² satırının tamamı da böyle.

Kullanım:  python dijital-baski-ekle.py [--uygula]
"""
import subprocess
import sys
import tempfile
import os
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
VT = r"C:\Users\Administrator\Desktop\Vinilturk-Fiyat-Listesi-2026-08-28.xlsx"
SSH = ["ssh", "-i", "/c/Users/Administrator/.ssh/markala_appvm", "root@10.20.33.5"]


def psql(sql: str) -> str:
    fd, yol = tempfile.mkstemp(suffix=".sql")
    with os.fdopen(fd, "w", encoding="utf-8") as f:
        f.write(sql)
    try:
        r = subprocess.run(["scp", "-i", "/c/Users/Administrator/.ssh/markala_appvm", "-q",
                            yol, "root@10.20.33.5:/tmp/_db.sql"], capture_output=True, text=True)
        if r.returncode != 0:
            raise SystemExit(f"scp: {r.stderr}")
        cmd = ("docker cp /tmp/_db.sql markala-postgres:/tmp/_db.sql && docker exec markala-postgres sh -c "
               "'psql -U $POSTGRES_USER -d $POSTGRES_DB -t -A -v ON_ERROR_STOP=1 -f /tmp/_db.sql'")
        r = subprocess.run(SSH + [cmd], capture_output=True, text=True, encoding="utf-8")
        if r.returncode != 0:
            raise SystemExit(f"psql: {r.stderr}")
        return r.stdout.strip()
    finally:
        os.unlink(yol)


# ── Viniltürk fiyatlarını oku ────────────────────────────────────────────────
wb = openpyxl.load_workbook(VT, data_only=True)
U = {}
for r in wb["Ürün Fiyatları"].iter_rows(min_row=2, values_only=True):
    if r[1]:
        U[str(r[1]).strip()] = {"satis": float(r[5] or 0), "maks": int(r[7] or 0)}
# Ek seçenek adları BENZERSİZ DEĞİL: "İç Mekan Baskı" hem Folyo (2 $, direkt) hem
# Kağıt (1 $, çevre ile çarpılır) grubunda var. Yalnız ada göre anahtarlarsak ikincisi
# birincisini eziyor ve yanlış fiyat giriliyor → (grup, ad) ile anahtarlanır.
E = {}
for r in wb["Ek Seçenek Fiyatları"].iter_rows(min_row=2, values_only=True):
    if r[4] and not str(r[4]).startswith("—"):
        E[(str(r[1]).strip(), str(r[4]).strip())] = {
            "fiyat": float(r[5] or 0), "kur": str(r[6] or ""), "etki": str(r[7] or "")}

vt = lambda ad: U[ad]["satis"]
mx = lambda ad: U[ad]["maks"]

sql = ["BEGIN;", "-- Viniltürk 28.08.2026 listesinden üretildi (scripts/katalog/dijital-baski-ekle.py)"]
ozet = []


def opsiyon(slug, gkey, glabel, gsort, key, label, sub, sort, cost, maks=0, effect="perM2"):
    kural = f'{{"effect":"{effect}","birim":"dolar"' + (f',"maxM2":{maks}' if maks else "") + "}"
    sql.append(f"""
INSERT INTO product_options (id, product_id, group_key, group_label, group_role, group_sort, option_key, option_label, option_sublabel, option_sort, locked, rules)
SELECT gen_random_uuid()::text, p.id, '{gkey}', '{glabel}', 'priced', {gsort}, '{key}', '{label}', {'NULL' if sub is None else "'" + sub + "'"}, {sort}, false, '{kural}'::jsonb
FROM products p WHERE p.slug = '{slug}'
  AND NOT EXISTS (SELECT 1 FROM product_options x WHERE x.product_id=p.id AND x.group_key='{gkey}' AND x.option_key='{key}');
INSERT INTO product_prices (id, product_id, group_key, option_key, dim_key, cost, price, created_at, updated_at)
SELECT gen_random_uuid()::text, p.id, '{gkey}', '{key}', NULL, {cost}, 0, now(), now()
FROM products p WHERE p.slug = '{slug}'
  AND NOT EXISTS (SELECT 1 FROM product_prices x WHERE x.product_id=p.id AND x.group_key='{gkey}' AND x.option_key='{key}');""")
    ozet.append((slug, f"{glabel}/{label}", cost))


# ── A) Mevcut ürünlere eksik malzemeler ──────────────────────────────────────
FOLYO_EK = [
    ("arkasi-gri-folyo",     "Arkası Gri Folyo",      "ışık geçirmez, arka yüz gri",       "Arkası Gri Folyo",        20),
    ("arkasi-gri-mat-folyo", "Arkası Gri Mat Folyo",  "mat yüzey, ışık geçirmez",          "Arkası Gri Mat Folyo",    21),
    ("laminasyonlu-folyo",   "Laminasyonlu Folyo",    "çizilmeye karşı korumalı",          "Laminasyonlu Folyo",      22),
    ("reflektif-folyo",      "Reflektif Folyo",       "solvent · ışık yansıtır",           "Reflektif Folyo Solvent Baskı", 23),
    ("reflektif-folyo-uv",   "Reflektif Folyo — UV",  "UV baskı · ışık yansıtır",          "Reflektif Folyo UV Baskı",24),
    ("lumen-folyo",          "Lümen Folyo",           "karanlıkta parlar (fotolüminesan)", "Lümen Folyo UV Baskı",    25),
]
for key, label, sub, vtad, sort in FOLYO_EK:
    opsiyon("folyo-cesitleri", "malzeme", "Malzeme", 0, key, label, sub, sort, vt(vtad), mx(vtad))

VINIL_EK = [
    ("avrupa-510gr-uv",  "Avrupa 510gr UV",        "kalın · UV baskı",        "Avrupa 510 Gr. UV Baskı",      20),
    ("arkasi-siyah-uv",  "Arkası Siyah UV",        "blockout · UV baskı",     "Arkası Siyah Vinil UV Baskı",  21),
    ("isikli-avrupa-uv", "Işıklı Avrupa UV",       "backlit · UV baskı",      "Işıklı Avrupa UV Baskı",       22),
    ("reflektif-vinil-uv","Reflektif Vinil UV",    "ışık yansıtır · UV baskı","Reflektif Vinil UV Baskı",     23),
]
for key, label, sub, vtad, sort in VINIL_EK:
    opsiyon("vinil-branda-440gr", "malzeme", "Malzeme", 0, key, label, sub, sort, vt(vtad), mx(vtad))

# One Way Vision artık AYRI ürün → Folyo Çeşitleri'nden çıkarılır (çakışma olmasın).
sql.append("""
DELETE FROM product_prices WHERE product_id=(SELECT id FROM products WHERE slug='folyo-cesitleri') AND option_key='one-way-vision';
DELETE FROM product_options WHERE product_id=(SELECT id FROM products WHERE slug='folyo-cesitleri') AND option_key='one-way-vision';""")

# ── B) Yeni ürünler ──────────────────────────────────────────────────────────
YENI = [
    dict(slug="baskes-folyo", ad="Baskes Folyo — Baskılı Kesim Folyo", kat="folyo",
         kisa="Baskı + kontür kesim tek işlemde — logo ve yazı formunda folyo",
         boyut="m² hesabı · UV baskı · kontür kesim",
         aciklama="Baskes, folyonun basılıp aynı anda şekline göre kesilmesidir; logo, yazı ve özel formlar arka fon olmadan tek parça çıkar. Vitrin yazısı, araç üstü logo ve tabela uygulamalarında kullanılır. Uygulama bandı (transfer tape) ile bütün hâlinde tek seferde yapıştırılır. 7 cm'den küçük parçalarda ek kesim ücreti uygulanır.",
         opsiyonlar=[("folyo-baskes","Folyo Baskes","standart beyaz folyo","Folyo Baskes"),
                     ("mat-folyo-baskes","Mat Folyo Baskes","parlama yapmaz","Mat Folyo Baskes"),
                     ("arkasi-gri-baskes","Arkası Gri Folyo Baskes","ışık geçirmez","Arkası Gri Folyo Baskes"),
                     ("arkasi-gri-mat-baskes","Arkası Gri Mat Baskes","mat · ışık geçirmez","Arkası Gri Mat Folyo Baskes"),
                     ("seffaf-baskes","Şeffaf Folyo Baskes","cam üstü uygulama","Şeffaf Folyo Baskes"),
                     ("kumlama-baskes","Kumlama Baskes","buzlu cam görünümü","Kumlama Baskes"),
                     ("kumlama-kesim","Kumlama — Sadece Kesim","baskısız, yalnız kesim","Kumlama SADECE KESİM - BASKISIZ")]),
    dict(slug="duvar-kagidi-baski", ad="Duvar Kağıdı Baskı — Özel Tasarım", kat="folyo",
         kisa="Ölçüye özel duvar kağıdı, solvent veya UV baskı",
         boyut="m² hesabı · solvent / UV",
         aciklama="Ofis, mağaza, kafe ve ev duvarları için ölçüye özel baskılı duvar kağıdı. İstediğiniz görsel, desen veya kurumsal tasarım duvarın tam ölçüsünde üretilir. Solvent baskı ekonomik ve dış mekana dayanıklıdır; UV baskı kokusuz olduğu için iç mekanda tercih edilir.",
         opsiyonlar=[("solvent","Solvent Baskı","ekonomik · dayanıklı","Duvar Kağıdı Solvent Baskı"),
                     ("uv","UV Baskı","kokusuz · iç mekan","Duvar Kağıdı UV Baskı")]),
    dict(slug="pleksi-baski", ad="Pleksi Baskı — UV Baskılı Akrilik Levha", kat="dekota-baski",
         kisa="3 ve 5 mm pleksi levhaya UV baskı — beyaz, siyah, şeffaf",
         boyut="m² hesabı · 3-5 mm · UV baskı",
         aciklama="Akrilik (pleksiglas) levha üzerine doğrudan UV baskı. Cam görünümlü, kırılmaz ve hafiftir; resepsiyon logosu, yönlendirme tabelası, menü panosu ve dekoratif pano uygulamalarında kullanılır. 3 mm fiyat/performans dengesi için, 5 mm daha dayanıklı ve prestijli işler için tercih edilir. CNC kesim ile istediğiniz forma getirilebilir.",
         opsiyonlar=None),
    dict(slug="kompozit-baski", ad="Kompozit Baskı — 3 mm Alüminyum Kompozit", kat="dekota-baski",
         kisa="3 mm alüminyum kompozit levhaya UV baskı",
         boyut="m² hesabı · 3 mm · UV baskı",
         aciklama="İki alüminyum tabaka arasında polietilen öz bulunan kompozit levhaya UV baskı. Dekotadan daha sert ve dış mekana daha dayanıklıdır; bina cephesi, kalıcı tabela ve yönlendirme panolarında kullanılır. Eğrilmez, paslanmaz. CNC kesim ile özel form verilebilir.",
         opsiyonlar=[("3mm","3 mm Kompozit","alüminyum kompozit levha","3 mm Kompozit Baskı")]),
    dict(slug="kanvas-tablo-baski", ad="Kanvas Tablo Baskı — Şasili Tuval", kat="dekota-baski",
         kisa="Tuval kumaşa baskı, solvent veya UV",
         boyut="m² hesabı · tuval · solvent / UV",
         aciklama="Fotoğraf ve tasarımlarınızın tuval (kanvas) kumaşa basılması. Ev, ofis, otel ve kafe duvarları için dokusu hissedilen mat bir yüzey verir. UV baskı kokusuz ve daha canlı renklidir; solvent baskı ekonomiktir.",
         opsiyonlar=[("solvent","Solvent Baskı","ekonomik","Canvas Solvent Baskı"),
                     ("uv","UV Baskı","kokusuz · canlı renk","Canvas UV Baskı")]),
    dict(slug="uv-dtf-baski", ad="UV DTF Baskı — Transfer Sticker", kat="folyo",
         kisa="Her yüzeye yapışan UV DTF transfer baskı, metraja göre fiyat",
         boyut="m² hesabı · metraja göre kademeli",
         aciklama="UV DTF, baskının transfer film üzerine yapılıp cam, metal, ahşap, plastik gibi neredeyse her yüzeye elle uygulanabildiği tekniktir. Kesim gerektirmez, dayanıklıdır ve kabartma dokusu verir. Fiyat toplam metraja göre kademelenir — çok metrede birim fiyat düşer.",
         opsiyonlar=[("0-2m","0 – 2 metre","az metrajlı işler","DTF 0 - 2 Metre"),
                     ("2-5m","2 – 5 metre","orta metraj","DTF 2 - 5 Metre"),
                     ("5-20m","5 – 20 metre","yüksek metraj","DTF 5 - 20 Metre"),
                     ("20m-uzeri","20 metre ve üzeri","en avantajlı kademe","DTF 20 Metre ve Üzeri")]),
    dict(slug="one-way-vision-baski", ad="One Way Vision Baskı — Tek Yön Görüş Folyosu", kat="folyo",
         kisa="Dışarıdan görsel, içeriden şeffaf delikli folyo",
         boyut="m² hesabı · solvent / UV",
         aciklama="Delikli yapısı sayesinde dışarıdan bakıldığında baskı görünen, içeriden bakıldığında camı şeffaf bırakan folyo. Mağaza vitrini, araç camı ve ofis camlarında kullanılır; gün ışığını kesmeden reklam alanı kazandırır.",
         opsiyonlar=[("solvent","Solvent Baskı","ekonomik","One Way Vision Solvent Baskı"),
                     ("uv","UV Baskı","kokusuz","One Way Vision UV Baskı")]),
]

for y in YENI:
    sql.append(f"""
INSERT INTO products (id, slug, name, category_id, short_description, description, base_price, starting_price,
                      production_time, size_label, images, badges, bestseller, is_active, parameters, pricing_mode, created_at, updated_at)
SELECT gen_random_uuid()::text, '{y["slug"]}', '{y["ad"].replace("'","''")}',
       (SELECT id FROM categories WHERE slug='{y["kat"]}'),
       '{y["kisa"].replace("'","''")}', '{y["aciklama"].replace("'","''")}',
       0, NULL, '2-3 iş günü', '{y["boyut"]}', ARRAY[]::text[], ARRAY['yeni']::text[],
       false, true, '[]'::jsonb, 'area', now(), now()
WHERE NOT EXISTS (SELECT 1 FROM products WHERE slug='{y["slug"]}');""")

    if y["slug"] == "pleksi-baski":
        # Kalınlık × Renk tek grup: renk fiyatı değiştirmiyor, kalınlık değiştiriyor.
        for i, (kal, vtad) in enumerate([("3 mm", "3 mm"), ("5 mm", "5 mm")]):
            for j, renk in enumerate(["Beyaz", "Siyah", "Şeffaf"]):
                anahtar = f"{kal.replace(' ','')}-{renk.lower().replace('ş','s').replace('ı','i')}"
                opsiyon(y["slug"], "malzeme", "Kalınlık × Renk", 0, anahtar,
                        f"{kal} — {renk}", None, i * 3 + j, E[("Kalınlık Pleksi Baskı", vtad)]["fiyat"], mx("Pleksi Baskı"))
    else:
        for i, (key, label, sub, vtad) in enumerate(y["opsiyonlar"]):
            opsiyon(y["slug"], "malzeme", "Malzeme", 0, key, label, sub, i, vt(vtad), mx(vtad))

# ── C) Ek işlemler ───────────────────────────────────────────────────────────
# VT "Fiyata Etkisi" → motor effect eşlemesi:
#   "Direkt eklenir"              → perPiece      (tl × adet)
#   "m² ile çarpılıp eklenir"     → perM2
#   "Çevre (m) ile çarpılıp ekl." → perPerimeter
EK_ISLEM = {
    "vinil-branda-440gr": [("dikis-kopca", "Dikiş + Kopça", "1 m² altı işlerde ücretli", ("Vinil Ek Seçenekler", "1 m2’den küçük işler dikiş + kopça"), "perPiece"),
                           ("kolon-dikis", "Kolon + Dikiş", "metre başına", ("Vinil Ek Seçenekler", "Kolon Dikiş (m)"), "perPerimeter"),
                           ("germe", "Germe", "ücretsiz", ("Vinil Ek Seçenekler", "Germe"), "perPiece")],
    "mesh-branda":        [("dikis-kopca", "Dikiş + Kopça", "1 m² altı işlerde ücretli", ("Vinil Ek Seçenekler", "1 m2’den küçük işler dikiş + kopça"), "perPiece"),
                           ("kolon-dikis", "Kolon + Dikiş", "metre başına", ("Vinil Ek Seçenekler", "Kolon Dikiş (m)"), "perPerimeter"),
                           ("germe", "Germe", "ücretsiz", ("Vinil Ek Seçenekler", "Germe"), "perPiece")],
    "folyo-cesitleri":    [("laminasyon", "Laminasyon", "çizilmeye karşı koruma", ("Folyo Ek Seçenekler", "Laminasyon"), "perPiece"),
                           ("ic-mekan", "İç Mekan Baskı", "kokusuz baskı", ("Folyo Ek Seçenekler", "İç Mekan Baskı"), "perPiece")],
    "dekota-baski-5mm":   [("cnc-kesim", "CNC Kesim", "özel forma kesim", ("Sert Zemin ve UV Seçenekler", "CNC Kesim"), "perM2")],
    "pleksi-baski":       [("cnc-kesim", "CNC Kesim", "özel forma kesim", ("Sert Zemin ve UV Seçenekler", "CNC Kesim"), "perM2")],
    "kompozit-baski":     [("cnc-kesim", "CNC Kesim", "özel forma kesim", ("Sert Zemin ve UV Seçenekler", "CNC Kesim"), "perM2")],
}
for slug, ekler in EK_ISLEM.items():
    # "Yok" seçeneği ücretsiz — fiyat satırı olmayan seçenek motorda 0 ekler.
    sql.append(f"""
INSERT INTO product_options (id, product_id, group_key, group_label, group_role, group_sort, option_key, option_label, option_sublabel, option_sort, locked)
SELECT gen_random_uuid()::text, p.id, 'ekislem', 'Ek İşlem', 'priced', 5, 'yok', 'Yok', 'ek işlem istemiyorum', 0, false
FROM products p WHERE p.slug='{slug}'
  AND NOT EXISTS (SELECT 1 FROM product_options x WHERE x.product_id=p.id AND x.group_key='ekislem' AND x.option_key='yok');""")
    for i, (key, label, sub, vtad, effect) in enumerate(ekler, start=1):
        opsiyon(slug, "ekislem", "Ek İşlem", 5, key, label, sub, i, E[vtad]["fiyat"], 0, effect)

sql.append("COMMIT;")
metin = "\n".join(sql)

print(f"{len(YENI)} yeni ürün · {len(ozet)} fiyatlı seçenek satırı\n")
son = None
for slug, ad, cost in ozet:
    if slug != son:
        print(f"■ {slug}")
        son = slug
    print(f"    {ad:46} {cost:>7.2f} $")

if "--uygula" not in sys.argv:
    print("\n(kuru çalışma — yazılmadı; --uygula ile çalıştır)")
    raise SystemExit
psql(metin)
print("\nUYGULANDI")
