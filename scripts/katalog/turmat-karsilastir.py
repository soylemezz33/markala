# -*- coding: utf-8 -*-
"""
TURMAT ↔ SİTE MALİYET KARŞILAŞTIRMASI (2026-08-29, Hasan: "tüm ürünlerin maliyeti
Turmatsan'la aynı olmalı, 0 hata").

Girdi: turmat-parsed.txt (turmat-ayikla.py) + fiyat-inceleme.txt (canlı DB dökümü).
Çıktı: markala-turmat-maliyet.xlsx — her Turmat kalemi bizdeki maliyetle yan yana.

0-HATA İLKESİ: otomatik eşleşme yalnız (bölüm→slug haritası) + (adet birebir) +
(açıklama-seçenek kelime kesişimi eşik üstü) sağlanırsa yapılır; aksi hâlde satır
"ELLE KONTROL" işaretlenir — yanlış eşleştirip yanlış maliyet basmaktansa insan bakar.
"Her X adet için Y" artışları DB'deki kademelerle açılır (base + n×artış).
"""
import re
import sys
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding="utf-8")
TURMAT = r"C:\Users\ADMINI~1\AppData\Local\Temp\turmat-parsed.txt"
DBDUMP = r"C:\Users\ADMINI~1\AppData\Local\Temp\fiyat-inceleme.txt"
HEDEF = r"C:\Users\Administrator\Downloads\markala-turmat-maliyet.xlsx"

# Turmat bölümü → site slug'ları (birden çok olabilir; İSG/area ürünleri Turmat DEĞİL).
BOLUM_SLUG = {
    "Kartvizit": ["klasik-kartvizit"],
    "Broşür": ["brosur"], "Pro Broşür": ["brosur"], "Selefonlu Broşür": ["brosur"],
    "El ilanı": ["el-ilani"],
    "Afişler": ["afis-105gr", "afis-135gr", "afis-170gr"],
    "Antetli": ["antetli-kagit"],
    "Zarf": ["zarf-diplomat-renkli", "zarf-diplomat-beyaz", "zarf-torba"],
    "Magnet": ["magnet", "arac-magneti"],
    "Amerikan Servis": ["amerikan-servis"],
    "KAPI ASKI BROŞÜRLERİ": ["kapi-aski-brosur"],
    "Dosyalar": ["cepli-dosya"],
    "Etiket": ["etiket"],
    "Makbuz": ["makbuz"],
    "Oto Paspas": ["oto-paspas"],
    "Küp Bloknot": ["kup-bloknot"],
    "Spiralli Bloknot": ["spiralli-bloknot"],
    "Kapaklı Bloknot": ["kapakli-bloknot"],
    "Kapaksız Bloknot": ["kapaksiz-bloknot"],
    "NOTLUK": ["notluk"],
    "ÇANTALAR": ["canta"],
    "ÜRÜN KUTULARI": ["urun-kutusu"],
}

def katla(s):
    tr = str.maketrans("ÇĞİÖŞÜI", "çğiöşüı")
    return s.translate(tr).lower()

def kelimeler(s):
    s = katla(s)
    s = re.sub(r"[^a-zçğıöşü0-9.,x]+", " ", s)
    return set(w for w in s.split() if len(w) >= 2)

# ── DB dökümü: slug → satırlar ──────────────────────────────────────────────
db = defaultdict(list)
ALAN = ["kat", "urun", "slug", "mod", "marj", "grup", "secenek",
        "gkey", "okey", "dim", "birim", "etki", "maliyet", "satis"]
for ham in open(DBDUMP, encoding="utf-8"):
    p = ham.rstrip("\n").split("§")
    if len(p) != len(ALAN):
        continue
    r = dict(zip(ALAN, p))
    if r["mod"] == "area":
        continue  # m² ürünleri Viniltürk kaynaklı — Turmat karşılaştırmasına girmez
    db[r["slug"]].append(r)

# ── Turmat kalemleri ────────────────────────────────────────────────────────
kalemler = []
for ham in open(TURMAT, encoding="utf-8"):
    p = ham.rstrip("\n").split("§")
    if len(p) != 6:
        continue
    bolum, kod, fiyat, acik, artis_adet, artis_fiyat = p
    kalemler.append({
        "bolum": bolum, "kod": kod, "fiyat": int(fiyat), "acik": acik,
        "artis_adet": int(artis_adet) if artis_adet else None,
        "artis_fiyat": int(artis_fiyat) if artis_fiyat else None,
    })

# Açıklamadan taban adet: "1000 Adet" / kod önekindeki 1/2/5/10 çarpanı vb.
def taban_adet(k):
    m = re.search(r"([0-9][0-9.]{2,6})\s*Adet", k["acik"], re.I)
    if m:
        return int(m.group(1).replace(".", ""))
    m = re.match(r"([0-9]{1,2})(CA|C|)[A-Z]", k["kod"])  # 1CA7=1000, 2CA7=2000, 5..=5000, 10..=10000
    if m and m.group(1) in ("1", "2", "5", "10"):
        return int(m.group(1)) * 1000
    if k["artis_adet"]:
        return k["artis_adet"]  # "2000 adet ... her 2000 için" düzeni
    return None

# ── Eşleştir + Excel ────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "KARŞILAŞTIRMA"
ws.append(["BÖLÜM", "KOD", "TURMAT AÇIKLAMA", "ADET", "TURMAT MALİYET",
           "EŞLEŞEN ÜRÜN", "SEÇENEK", "KADEME", "BİZDEKİ MALİYET", "FARK", "DURUM"])

ayni = farkli = elle = 0
eslesen_db = set()
for k in kalemler:
    sluglar = BOLUM_SLUG.get(k["bolum"], [])
    adaylar = [r for s in sluglar for r in db.get(s, [])]
    adet = taban_adet(k)
    kw = kelimeler(k["acik"] + " " + k["kod"])
    # Kademeleri aç: base + n×artış (DB'de o üründe hangi kademeler varsa)
    hedefler = []
    if adet and k["artis_adet"] and k["artis_fiyat"]:
        for r in adaylar:
            try:
                d = int(r["dim"])
            except ValueError:
                continue
            n = (d - adet) / k["artis_adet"]
            if n >= 0 and n == int(n):
                hedefler.append((r, k["fiyat"] + int(n) * k["artis_fiyat"]))
    elif adet:
        hedefler = [(r, k["fiyat"]) for r in adaylar if r["dim"] == str(adet)]
    else:
        hedefler = [(r, k["fiyat"]) for r in adaylar if not r["dim"]]

    # Seçenek metni kesişimiyle daralt (ör. "31x44" / "90 gr" / "mat selefon")
    def skor(r):
        return len(kw & kelimeler(r["secenek"] + " " + r["grup"] + " " + r["okey"]))
    if len({(r["okey"]) for r, _ in hedefler}) > 1:
        enyuksek = max((skor(r) for r, _ in hedefler), default=0)
        if enyuksek >= 1:
            hedefler = [(r, f) for r, f in hedefler if skor(r) == enyuksek]

    if not hedefler:
        ws.append([k["bolum"], k["kod"], k["acik"][:80], adet, k["fiyat"],
                   "", "", "", None, None, "ELLE KONTROL — eşleşme yok"])
        elle += 1
        continue
    if len({r["okey"] for r, _ in hedefler}) > 1:
        ws.append([k["bolum"], k["kod"], k["acik"][:80], adet, k["fiyat"],
                   "/".join(sorted({r["slug"] for r, _ in hedefler})), "BELİRSİZ", "", None, None,
                   "ELLE KONTROL — çok aday"])
        elle += 1
        continue
    for r, beklenen in sorted(hedefler, key=lambda x: int(x[0]["dim"] or 0)):
        mevcut = float(r["maliyet"]) if r["maliyet"] else None
        fark = None if mevcut is None else round(beklenen - mevcut, 2)
        durum = "✓ aynı" if fark == 0 else ("MALİYET BOŞ" if mevcut is None else "✗ FARKLI")
        if fark == 0:
            ayni += 1
        else:
            farkli += 1
        eslesen_db.add((r["slug"], r["okey"], r["dim"]))
        ws.append([k["bolum"], k["kod"], k["acik"][:80], r["dim"] or adet, beklenen,
                   r["slug"], r["secenek"], r["dim"], mevcut, fark, durum])

# Bizde olup Turmat'ta eşleşmeyen matbaa satırları (ters yön — tam kapsama)
ws2 = wb.create_sheet("TURMATTA EŞLEŞMEYEN")
ws2.append(["ÜRÜN", "SEÇENEK", "KADEME", "BİZDEKİ MALİYET",
            "NOT: Turmat kaynaklı olmayabilir (Viniltürk/elle) — bilgi amaçlı"])
turmat_sluglar = {s for v in BOLUM_SLUG.values() for s in v}
for slug in sorted(turmat_sluglar):
    for r in db.get(slug, []):
        if (r["slug"], r["okey"], r["dim"]) not in eslesen_db:
            ws2.append([r["slug"], r["secenek"], r["dim"], r["maliyet"]])

for sayfa in (ws, ws2):
    for c in sayfa[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4B3AA0")
    sayfa.freeze_panes = "A2"
kirmizi = PatternFill("solid", fgColor="FBE9E7")
for i in range(2, ws.max_row + 1):
    if str(ws.cell(row=i, column=11).value or "").startswith(("✗", "ELLE", "MALİYET")):
        for c in range(1, 12):
            ws.cell(row=i, column=c).fill = kirmizi
for col, wd in zip("ABCDEFGHIJK", [18, 12, 44, 8, 13, 20, 26, 8, 13, 10, 26]):
    ws.column_dimensions[col].width = wd

wb.save(HEDEF)
print(f"✓ aynı: {ayni} · ✗ farklı: {farkli} · elle kontrol: {elle}")
print(f"Turmat'ta eşleşmeyen bizim satır: {ws2.max_row - 1}")
print(f"→ {HEDEF}")
